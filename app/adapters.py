"""Dataset adapters for Geostat published spreadsheets.

Every adapter implements the same four-stage contract:

    discover()   -> SourceRef      where the file lives and what page documents it
    download()   -> bytes          raw HTTP fetch (browser User-Agent is mandatory)
    parse(bytes) -> list[dict]     workbook grid -> tidy records, markers preserved
    normalise()  -> list[Row]      long format with the composite key

`validate()` lives in contracts.py so the checks can run against any row set
(a fresh download, a committed vintage, or a deliberately corrupted copy).

The single most important thing in this file is `currency_for_year`. Geostat's
annual earnings workbook carries a footnote:

    "*Currency - before 1993 -Rouble; 1993 - Coupon; 1994 - Thousand Coupon;
     since 1995 - Lari."

The columns therefore are NOT one comparable series. 1993 reads 27950 and 1995
reads 13.5; that is a currency reform, not a 99.95% wage collapse. Every parsed
row carries the unit of its own year so the rest of the system cannot forget.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field, asdict
from typing import Callable

import openpyxl

from . import geography

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
)
WAGES_PAGE = "https://www.geostat.ge/en/modules/categories/39/wages"
INFLATION_PAGE = (
    "https://www.geostat.ge/en/modules/categories/26/inflation-consumer-price-index"
)
EMPLOYMENT_PAGE = (
    "https://www.geostat.ge/en/modules/categories/683/dasakmeba-umushevroba"
)
HOUSEHOLD_INCOME_PAGE = (
    "https://www.geostat.ge/en/modules/categories/50/households-income"
)
HOUSEHOLD_EXPENDITURE_PAGE = (
    "https://www.geostat.ge/en/modules/categories/51/households-expenditures"
)
POPULATION_PAGE = "https://www.geostat.ge/en/modules/categories/41/population"
BUSINESS_DEMOGRAPHY_PAGE = (
    "https://www.geostat.ge/en/modules/categories/69/business-demography"
)
REGIONAL_PAGE = (
    "https://www.geostat.ge/en/modules/categories/93/regional-statistics"
)

# Cells Geostat uses to mean "not published / not available". `NULL` is written
# as a literal string in the business demography workbook, where an employment
# figure is suppressed for disclosure control rather than being zero.
MISSING_MARKERS = {
    "…", "...", "..", "-", "—", "–", "", "n/a", "N/A", ":", "NULL", "null",
}

ROMAN_MONTHS = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
    "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12,
}

ROMAN_QUARTERS = {"I": 1, "II": 2, "III": 3, "IV": 4}


# --------------------------------------------------------------------------
# currency eras - the trap this whole project is built around
# --------------------------------------------------------------------------

CURRENCY_ERAS = [
    (None, 1992, "RUB", "Rouble"),
    (1993, 1993, "KUP", "Coupon"),
    (1994, 1994, "TKUP", "Thousand Coupon"),
    (1995, None, "GEL", "Lari"),
]

GEL_ERA_START = 1995


# --------------------------------------------------------------------------
# series breaks - the currency trap's cousin
# --------------------------------------------------------------------------

# A definition change is a discontinuity that no value check can see. Nothing
# is out of range, no unit moved, no period is missing, and the step is far too
# small to trip a 10x jump test - it just means the two halves of the line are
# counting different people.
#
# Geostat adopted the ICLS-19 employment standard in 2020, which reclassifies
# subsistence agricultural producers as outside employment, and recalculated
# 2010 onward on that basis. 1998-2009 were left on the old basis. The result
# is a break between 2009 and 2010: employment falls from 1,611 to 1,168
# thousand and the unemployment rate rises from 18.3% to 27.2% in one year,
# because the definition moved, not because 443,000 people lost their jobs.
SERIES_BREAKS: dict[str, list[dict]] = {
    "labour_force": [{
        "before": "2009",
        "after": "2010",
        "what": "ICLS-19 employment definition",
        "why": (
            "Subsistence agricultural producers count as employed up to 2009 "
            "and as outside employment from 2010. Geostat adopted the standard "
            "in 2020 and recalculated 2010-2019; 1998-2009 were left on the "
            "old basis. Employment falls 1,611 to 1,168 thousand and "
            "unemployment rises 18.3% to 27.2% across that single boundary "
            "because the question changed, not the labour market."
        ),
    }],
    "labour_force_by_region": [{
        "before": "2009",
        "after": "2010",
        "what": "ICLS-19 employment definition",
        "why": (
            "The same reclassification as the national series, and the same "
            "boundary. The regional file additionally stops publishing the "
            "hired / self-employed split for 2010-2019."
        ),
    }],
}


def series_breaks(dataset_id: str) -> list[dict]:
    return SERIES_BREAKS.get(dataset_id, [])


def spans_a_break(dataset_id: str, first: str, last: str) -> dict | None:
    """The break a first-to-last comparison would silently cross, if any."""
    for brk in series_breaks(dataset_id):
        if first <= brk["before"] and last >= brk["after"]:
            return brk
    return None


def currency_for_year(year: int) -> str:
    """Monetary unit in force in Georgia in `year`, per the Geostat footnote."""
    if year < 1993:
        return "RUB"
    if year == 1993:
        return "KUP"
    if year == 1994:
        return "TKUP"
    return "GEL"


def currency_name(code: str) -> str:
    for _lo, _hi, c, name in CURRENCY_ERAS:
        if c == code:
            return name
    return code


# --------------------------------------------------------------------------
# row model
# --------------------------------------------------------------------------

@dataclass
class Row:
    """One observation in long format.

    Composite key: (dataset_id, indicator_code, breakdown_code, period, unit,
    vintage_id). `vintage_id` is stamped when the row is written into a vintage.
    """

    dataset_id: str
    indicator_code: str
    breakdown_code: str
    breakdown_label: str
    period: str                 # "1995", "1995-03" or "1995-Q1"
    unit: str
    value: float | None
    raw: str                    # exactly what the cell held
    status: str                 # ok | missing | unparsed
    is_preliminary: bool = False
    vintage_id: str = ""
    # The printed name of the measure. Empty on the datasets where the measure
    # is fixed and the row label is the breakdown (earnings by region, and so
    # on); populated where the sheet varies the measure down the rows instead,
    # as the household and labour-force-by-region workbooks do. Kept out of the
    # required schema on purpose so vintages committed before it existed still
    # validate.
    indicator_label: str = ""

    def key(self) -> tuple:
        return (
            self.dataset_id, self.indicator_code, self.breakdown_code,
            self.period, self.unit, self.vintage_id,
        )

    def to_dict(self) -> dict:
        return asdict(self)


def parse_number(raw_value) -> tuple[float | None, str, str]:
    """Return (value, status, raw_text).

    status is 'ok', 'missing' (a published gap marker) or 'unparsed' (something
    we did not expect). Nothing is ever silently coerced to 0.
    """
    if raw_value is None:
        return None, "missing", ""
    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
        return float(raw_value), "ok", repr(raw_value)
    text = str(raw_value).strip()
    if text in MISSING_MARKERS:
        return None, "missing", text
    cleaned = text.replace("\xa0", "").replace(" ", "")
    # Geostat publishes decimal points. A comma here means someone handed us a
    # European-formatted export; we refuse to guess whether it is a decimal
    # separator or a thousands separator.
    if "," in cleaned:
        return None, "unparsed", text
    try:
        return float(cleaned), "ok", text
    except ValueError:
        return None, "unparsed", text


def parse_period_header(cell) -> tuple[str | None, bool]:
    """Turn a period header cell into (canonical period, is_preliminary).

    `2025**`   -> ("2025", True) because ** is the preliminary-data footnote.
    `2006*`    -> ("2006", False) because a single * is a methodology footnote.
    `2007_I`   -> ("2007-Q1")  the quarterly earnings workbook's own notation,
    `2008 IV`  -> ("2008-Q4")  which switches between an underscore and a space
                               partway across the same header row.

    Quarters are canonicalised to `YYYY-QN` so that a period sorts, compares
    and joins the same way whatever the workbook wrote.
    """
    if cell is None:
        return None, False
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        return str(int(cell)), False
    text = str(cell).strip()
    prelim = text.endswith("**")
    quarter = re.match(r"^(\d{4})\s*[_\-\s]\s*(I{1,3}|IV)\b", text, re.IGNORECASE)
    if quarter:
        roman = quarter.group(2).upper()
        return f"{quarter.group(1)}-Q{ROMAN_QUARTERS[roman]}", prelim
    digits = re.match(r"^(\d{4})", text)
    if not digits:
        return None, False
    return digits.group(1), prelim


def period_year(period: str) -> int:
    """Calendar year of any canonical period: `2007`, `2007-03`, `2007-Q1`."""
    return int(period[:4])


def measure_code(text: str) -> str:
    """Stable code for a measure printed as an enumerated sheet row.

    The household workbooks number their rows and state the identity in the
    label - `1. Income, total (2+3)`. The number and the formula are useful to
    a reader and useless in a key, so the code keeps the words only. The
    printed text survives on `Row.indicator_label`, and the identities the
    formulas express are checked by the IDENTITY contract.
    """
    cleaned = re.sub(r"^\s*\d+[.)]\s*", "", str(text).strip())
    cleaned = re.sub(r"\s*\([0-9+\-\s]*\)\s*$", "", cleaned)
    return slug(cleaned)


def slug(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", str(text).strip().lower())
    return s.strip("_") or "unknown"


# --------------------------------------------------------------------------
# parsers - five shapes cover all twelve published workbooks
# --------------------------------------------------------------------------

def _sheet(data: bytes, name: str):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    if name not in wb.sheetnames:
        raise KeyError(f"sheet {name!r} not in workbook (have {wb.sheetnames})")
    return wb[name]


def _last_col(ws, row: int) -> int:
    last = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value is not None:
            last = c
    return last


def _year_headers(ws, header_row: int) -> list[tuple[int, str, bool]]:
    out = []
    for c in range(2, _last_col(ws, header_row) + 1):
        period, prelim = parse_period_header(ws.cell(header_row, c).value)
        if period:
            out.append((c, period, prelim))
    return out


def parse_year_grid(
    data: bytes,
    *,
    sheet: str,
    header_row: int,
    first_row: int,
    unit_mode: str,
    indicator: str,
    dataset_id: str,
    breakdown_prefix: str = "",
    detect_sections: bool = False,
    label_axis: str = "breakdown",
) -> list[Row]:
    """Label column A, year headers across the top.

    `detect_sections=True` handles the annual earnings workbook, where rows like
    `Sex` / `Type of ownership` / `Sector` are group headings with no data and
    the rows beneath them belong to that group.

    `label_axis` says what the row label *is*. On the earnings sheets the
    measure is fixed and each row is a breakdown of it - a region, an activity.
    On the labour force sheet the opposite holds: every row is a different
    measure of one population. Putting the measure on the indicator axis in
    both cases is what lets the IDENTITY contract state `labour force =
    employed + unemployed` once and have it mean the same thing in the national
    file and the regional one.
    """
    ws = _sheet(data, sheet)
    headers = _year_headers(ws, header_row)
    rows: list[Row] = []
    section = ""
    for r in range(first_row, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()
        if not label or label.startswith("*") or label.lower().startswith("source"):
            continue
        cells = [(p, prelim, ws.cell(r, c).value) for c, p, prelim in headers]
        has_data = any(v is not None and str(v).strip() != "" for _, _, v in cells)
        if detect_sections and not has_data:
            section = slug(label)
            continue
        if not has_data:
            continue
        code_parts = [x for x in (breakdown_prefix, section, slug(label)) if x]
        code = ".".join(code_parts)
        if label_axis == "indicator":
            row_indicator, row_label = measure_code(label), label
            row_breakdown = indicator
            # `indicator` carries the breakdown code on this axis, so give it a
            # printed name: a CSV column reading `country.georgia` where the
            # label belongs is a code leaking into human-facing output.
            row_breakdown_label = (
                geography.display_name(indicator)
                if "." in indicator else indicator
            )
        else:
            row_indicator, row_label = indicator, ""
            row_breakdown, row_breakdown_label = code, label
        for period, prelim, cell in cells:
            value, status, raw = parse_number(cell)
            unit = (currency_for_year(period_year(period))
                    if unit_mode == "currency_era" else unit_mode)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=row_indicator,
                breakdown_code=row_breakdown, breakdown_label=row_breakdown_label,
                period=period, unit=unit, value=value, raw=raw,
                status=status, is_preliminary=prelim,
                indicator_label=row_label,
            ))
    return rows


def parse_banded_grid(
    data: bytes, *, sheet: str, header_row: int, first_row: int,
    unit_mode: str, indicator: str, dataset_id: str,
) -> list[Row]:
    """Same as a year grid, but split into bands (Women / Men) whose heading
    sits in column B with column A empty."""
    ws = _sheet(data, sheet)
    headers = _year_headers(ws, header_row)
    rows: list[Row] = []
    band = ""
    band_label = ""
    for r in range(first_row, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and isinstance(b, str) and b.strip():
            band, band_label = slug(b), b.strip()
            continue
        if a is None:
            continue
        label = str(a).strip()
        if not label or label.startswith("*"):
            continue
        for c, period, prelim in headers:
            value, status, raw = parse_number(ws.cell(r, c).value)
            unit = (currency_for_year(period_year(period))
                    if unit_mode == "currency_era" else unit_mode)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=f"{band}.{slug(label)}",
                breakdown_label=f"{band_label} - {label}",
                period=period, unit=unit, value=value, raw=raw,
                status=status, is_preliminary=prelim,
            ))
    return rows


def _year_band_rows(ws) -> list[tuple[int, str]]:
    """Rows whose column A is nothing but a four-digit year.

    Three of the regional workbooks are stacked one year-block per band rather
    than one year per column, so the year lives in the gutter.
    """
    out = []
    for r in range(1, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if value is None:
            continue
        text = str(value).strip()
        if re.fullmatch(r"\d{4}", text):
            out.append((r, text))
    return out


def parse_year_banded_matrix(
    data: bytes,
    *,
    sheet: str,
    dataset_id: str,
    unit_mode: str,
    fixed_header_row: int | None = None,
    percent_suffixes: tuple[str, ...] = (),
) -> list[Row]:
    """Year blocks stacked down the sheet; places across, measures down.

    Two published layouts share this shape and differ in one detail:

    * labour force by region repeats the region header row inside every year
      block, because the set of regions changes - 2003 publishes `The remaining
      regions`, later years enumerate all eleven;
    * household income and expenditure print the region header once at the top.

    `fixed_header_row` selects the second layout. Reading the header once for
    the first would attribute 2020's columns to 2003's region set.
    """
    ws = _sheet(data, sheet)
    bands = _year_band_rows(ws)
    if not bands:
        raise ValueError(f"sheet {sheet!r} has no year bands in column A")
    band_starts = [r for r, _ in bands]

    rows: list[Row] = []
    for index, (band_row, year) in enumerate(bands):
        stop = band_starts[index + 1] if index + 1 < len(bands) else ws.max_row + 1
        if fixed_header_row is not None:
            header_row, first_row = fixed_header_row, band_row + 1
        else:
            header_row, first_row = band_row + 1, band_row + 2

        places = []
        for c in range(2, _last_col(ws, header_row) + 1):
            label = ws.cell(header_row, c).value
            if isinstance(label, str) and label.strip():
                code, name = geography.resolve(label)
                places.append((c, code, name))
        if not places:
            continue

        for r in range(first_row, stop):
            measure = ws.cell(r, 1).value
            if measure is None:
                continue
            measure = str(measure).strip()
            if not measure or measure.startswith("*"):
                continue
            if measure.lower().startswith(("source", "note", "http", "last update")):
                continue
            # The household sheets put their footnotes in the measure column,
            # prefixed with the same ellipsis they use for a suppressed cell:
            # "… The data is not available due to small sample size…". A
            # sentence is not a measure, and letting one through would mint an
            # indicator code 90 characters long that joins to nothing.
            if measure.lstrip("… .").startswith(("The data", "Data")):
                continue
            if measure[:1] in {"…", ".", "*"}:
                continue
            indicator = measure_code(measure)
            unit = unit_mode
            if percent_suffixes and measure.rstrip(" .").endswith(percent_suffixes):
                unit = "percent"
            for col, code, name in places:
                value, status, raw = parse_number(ws.cell(r, col).value)
                rows.append(Row(
                    dataset_id=dataset_id, indicator_code=indicator,
                    breakdown_code=code, breakdown_label=name,
                    period=year, unit=unit, value=value, raw=raw, status=status,
                    indicator_label=measure,
                ))
    return rows


def parse_indicator_columns_by_year_band(
    data: bytes,
    *,
    sheet: str,
    dataset_id: str,
    header_row: int,
    label_col: int,
    first_col: int,
    units: dict[str, str],
) -> list[Row]:
    """Business demography: indicators across, places down, year in a band row.

    The year sits in the first data column of an otherwise empty row, so a band
    is detected by "label cell empty, first data cell is a year".
    """
    ws = _sheet(data, sheet)
    indicators = []
    for c in range(first_col, _last_col(ws, header_row) + 1):
        label = ws.cell(header_row, c).value
        if isinstance(label, str) and label.strip():
            clean = " ".join(str(label).split())
            indicators.append((c, slug(clean.rstrip("*")), clean))

    rows: list[Row] = []
    year = ""
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, label_col).value
        band = ws.cell(r, first_col).value
        if (label is None or not str(label).strip()) and band is not None:
            period, _prelim = parse_period_header(band)
            if period:
                year = period
            continue
        if label is None or not year:
            continue
        text = str(label).strip()
        if not text or text.startswith("*"):
            continue
        code, name = geography.resolve(text)
        for col, indicator, printed in indicators:
            value, status, raw = parse_number(ws.cell(r, col).value)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=code, breakdown_label=name,
                period=year, unit=units.get(indicator, "count"),
                value=value, raw=raw, status=status,
            ))
    return rows


def parse_geo_year_grid(
    data: bytes, *, sheet: str, header_row: int, first_row: int,
    dataset_id: str, indicator: str, unit: str,
) -> list[Row]:
    """Population: places down, years across, level carried by the cell indent.

    Geostat prints regions flush left and their municipalities indented under
    them in the same column. That indent is the only published signal of which
    level a row belongs to, and without it `Tbilisi` the region and `Tbilisi`
    the municipality are one row of numbers counted twice.
    """
    ws = _sheet(data, sheet)
    headers = _year_headers(ws, header_row)
    rows: list[Row] = []
    for r in range(first_row, ws.max_row + 1):
        cell = ws.cell(r, 1)
        label = cell.value
        if label is None:
            continue
        text = str(label).strip()
        if not text or text.startswith(("*", "-", "Note")):
            continue
        indented = (cell.alignment.indent or 0) > 0
        level = geography.MUNICIPALITY if indented else None
        # `C. Ozurgeti*` and friends are city rows whose values are folded into
        # the surrounding municipality; the sheet says so in its own footnote.
        code, name = geography.resolve(text.rstrip("*"), level=level)
        for col, period, prelim in headers:
            value, status, raw = parse_number(ws.cell(r, col).value)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=code, breakdown_label=name,
                period=period, unit=unit, value=value, raw=raw,
                status=status, is_preliminary=prelim,
            ))
    return rows


def parse_two_level_header(
    data: bytes, *, sheet: str, year_row: int, measure_row: int, first_row: int,
    dataset_id: str, unit: str, measure_codes: dict[str, str],
) -> list[Row]:
    """Gender pay gap: year on one header row, measure on the next.

    The year cell is merged across its measures, so it is carried forward from
    the left rather than read per column.
    """
    ws = _sheet(data, sheet)
    columns = []
    year = ""
    for c in range(2, ws.max_column + 1):
        period, _prelim = parse_period_header(ws.cell(year_row, c).value)
        if period:
            year = period
        measure = ws.cell(measure_row, c).value
        if not year or not isinstance(measure, str) or not measure.strip():
            continue
        key = slug(measure)
        code = next((v for k, v in measure_codes.items() if k in key), None)
        if code:
            columns.append((c, year, code))

    rows: list[Row] = []
    section = ""
    for r in range(first_row, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        text = " ".join(str(label).split())
        if not text or text.startswith("*"):
            continue
        cells = [(p, code, ws.cell(r, c).value) for c, p, code in columns]
        if not any(v is not None and str(v).strip() != "" for _p, _c, v in cells):
            # `by economic activity` / `by occupation` head a group of rows and
            # carry no numbers of their own.
            if not text.lower().startswith("of which"):
                section = slug(text.removeprefix("by "))
            continue
        breakdown = ".".join(x for x in (section, slug(text)) if x)
        for period, indicator, cell in cells:
            value, status, raw = parse_number(cell)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=breakdown, breakdown_label=text,
                period=period, unit=unit, value=value, raw=raw, status=status,
            ))
    return rows


def parse_quarter_rows(
    data: bytes, *, sheet: str, header_row: int, first_row: int,
    dataset_id: str, indicator: str, unit: str,
) -> list[Row]:
    """Tourism: year and roman-numeral quarter in the first two columns.

    The year is printed once per group and blank on the rows beneath it, so it
    is carried down. A blank year with no preceding year is a footnote, not a
    2014 observation.
    """
    ws = _sheet(data, sheet)
    places = []
    for c in range(3, _last_col(ws, header_row) + 1):
        label = ws.cell(header_row, c).value
        if isinstance(label, str) and label.strip():
            places.append((c, *geography.resolve(label)))

    rows: list[Row] = []
    year = ""
    for r in range(first_row, ws.max_row + 1):
        year_cell = ws.cell(r, 1).value
        quarter_cell = ws.cell(r, 2).value
        if year_cell is not None:
            text = str(year_cell).strip()
            if re.fullmatch(r"\d{4}", text):
                year = text
            else:
                # `Source:` / `Metadata:` / a bare URL ends the table.
                if not re.fullmatch(r"\s*", text):
                    break
        roman = str(quarter_cell or "").strip().upper()
        if not year or roman not in ROMAN_QUARTERS:
            continue
        period = f"{year}-Q{ROMAN_QUARTERS[roman]}"
        for col, code, name in places:
            value, status, raw = parse_number(ws.cell(r, col).value)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=code, breakdown_label=name,
                period=period, unit=unit, value=value, raw=raw, status=status,
            ))
    return rows


def parse_place_year_grid(
    data: bytes, *, sheet: str, header_row: int, first_row: int,
    dataset_id: str, indicator: str, unit: str,
) -> list[Row]:
    """Places down, years across, no indent levels (tourism hotels and guests)."""
    ws = _sheet(data, sheet)
    headers = _year_headers(ws, header_row)
    rows: list[Row] = []
    for r in range(first_row, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        text = " ".join(str(label).split())
        if not text or text.startswith("*"):
            continue
        if text.lower().startswith(
            ("source", "metadata", "http", "note", "last update")
        ):
            break
        code, name = geography.resolve(text)
        for col, period, prelim in headers:
            value, status, raw = parse_number(ws.cell(r, col).value)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code=indicator,
                breakdown_code=code, breakdown_label=name,
                period=period, unit=unit, value=value, raw=raw,
                status=status, is_preliminary=prelim,
            ))
    return rows


def parse_cpi_monthly(data: bytes, *, dataset_id: str) -> list[Row]:
    """CPI 2010=100. One sheet per city, years down, roman-numeral months across.

    Also derives an annual average per city, but only for years where all twelve
    months are published - a partial year would quietly understate the average.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows: list[Row] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        city = sheet_name.strip()
        code = slug(city)
        months = []
        for c in range(2, _last_col(ws, 3) + 1):
            key = str(ws.cell(3, c).value or "").strip().upper()
            if key in ROMAN_MONTHS:
                months.append((c, ROMAN_MONTHS[key]))
        for r in range(4, ws.max_row + 1):
            year_cell = ws.cell(r, 1).value
            if not isinstance(year_cell, (int, float)):
                continue
            year = int(year_cell)
            monthly: list[float] = []
            for c, m in months:
                value, status, raw = parse_number(ws.cell(r, c).value)
                if status == "missing" and raw == "":
                    continue  # month not yet published
                if value is not None:
                    monthly.append(value)
                rows.append(Row(
                    dataset_id=dataset_id, indicator_code="cpi_2010_100",
                    breakdown_code=code, breakdown_label=city,
                    period=f"{year}-{m:02d}", unit="index_2010_100",
                    value=value, raw=raw, status=status,
                ))
            if len(monthly) == 12:
                rows.append(Row(
                    dataset_id=dataset_id, indicator_code="cpi_annual_avg_2010_100",
                    breakdown_code=code, breakdown_label=city,
                    period=str(year), unit="index_2010_100",
                    value=round(sum(monthly) / 12, 6),
                    raw="derived: mean of 12 published months",
                    status="ok",
                ))
    return rows


def parse_cpi_yoy(data: bytes, *, dataset_id: str) -> list[Row]:
    """CPI vs the same month of the previous year.

    The workbook carries the full COICOP tree for seven cities across ~270
    columns. We keep the headline 'Total' line per city: that is the number the
    press releases quote, and it is the only line the rest of GeoStats consumes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows: list[Row] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        city = sheet_name.strip()
        code = slug(city)
        # row 3 holds a year every 12 columns, row 4 the roman months
        year_at: dict[int, int] = {}
        current = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            if isinstance(v, (int, float)) and 1990 < float(v) < 2100:
                current = int(v)
            month_key = str(ws.cell(4, c).value or "").strip().upper()
            if current and month_key in ROMAN_MONTHS:
                year_at[c] = current
        total_row = None
        for r in range(4, min(ws.max_row, 12) + 1):
            if str(ws.cell(r, 3).value or "").strip().lower() == "total":
                total_row = r
                break
        if total_row is None:
            continue
        for c, year in year_at.items():
            month = ROMAN_MONTHS[str(ws.cell(4, c).value).strip().upper()]
            value, status, raw = parse_number(ws.cell(total_row, c).value)
            if status == "missing" and raw == "":
                continue
            rows.append(Row(
                dataset_id=dataset_id, indicator_code="cpi_same_month_prev_year",
                breakdown_code=code, breakdown_label=city,
                period=f"{year}-{month:02d}", unit="index_prev_year_100",
                value=value, raw=raw, status=status,
            ))
    return rows


def parse_basket_weights(data: bytes, *, dataset_id: str) -> list[Row]:
    """Consumer basket weights. Header spans two rows: labels on row 3, years on
    row 4 starting at column E. Values are shares (Total = 1.0), not percent,
    despite the title saying '%'."""
    ws = _sheet(data, "Weights")
    headers = []
    for c in range(5, _last_col(ws, 4) + 1):
        period, prelim = parse_period_header(ws.cell(4, c).value)
        if period:
            headers.append((c, period, prelim))
    rows: list[Row] = []
    for r in range(5, ws.max_row + 1):
        level = ws.cell(r, 2).value
        coicop = ws.cell(r, 3).value
        label = ws.cell(r, 4).value
        if coicop is None and label is None:
            continue
        name = str(label).strip() if label else str(coicop).strip()
        # The COICOP code alone is NOT unique in this workbook: "11" is Food at
        # level 3 and Restaurants and hotels at level 2. The key-uniqueness
        # contract caught that; the level has to be part of the key.
        code = slug(str(coicop) if coicop is not None else name)
        if level is not None:
            code = f"l{int(level)}.{code}"
        for c, period, prelim in headers:
            value, status, raw = parse_number(ws.cell(r, c).value)
            rows.append(Row(
                dataset_id=dataset_id, indicator_code="basket_weight",
                breakdown_code=code, breakdown_label=name,
                period=period, unit="share_of_1", value=value, raw=raw,
                status=status, is_preliminary=prelim,
            ))
    return rows


# --------------------------------------------------------------------------
# adapter registry
# --------------------------------------------------------------------------

@dataclass
class Adapter:
    dataset_id: str
    title: str
    url: str
    source_page: str
    parser: Callable[[bytes], list[Row]]
    unit_family: str
    note: str
    expected_sheets: list[str] = field(default_factory=list)
    period_grain: str = "annual"

    # -- the four ingestion stages ------------------------------------------
    def discover(self) -> dict:
        """Where this dataset comes from. Geostat has no machine-readable
        release feed on these pages, so the file URL is pinned and the vintage
        log is what proves whether the bytes behind it changed."""
        return {
            "dataset_id": self.dataset_id,
            "url": self.url,
            "source_page": self.source_page,
            "expected_sheets": self.expected_sheets,
        }

    def download(self, timeout: float = 45.0) -> tuple[bytes, int]:
        """Fetch the workbook. The browser User-Agent is not optional: without
        it Geostat answers 200 with a zero-byte body."""
        import httpx

        headers = {
            "User-Agent": USER_AGENT,
            "Referer": self.source_page,
            "Accept": "*/*",
        }
        with httpx.Client(follow_redirects=True, timeout=timeout) as client:
            resp = client.get(self.url, headers=headers)
        return resp.content, resp.status_code

    def parse(self, data: bytes) -> list[Row]:
        return self.parser(data)

    def normalise(self, data: bytes, vintage_id: str) -> list[Row]:
        rows = self.parse(data)
        for row in rows:
            row.vintage_id = vintage_id
        return rows


def _earnings_annual(data: bytes) -> list[Row]:
    return parse_year_grid(
        data, sheet="1", header_row=3, first_row=4,
        unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
        dataset_id="earnings_annual", detect_sections=True,
    )


def _median_earnings(data: bytes) -> list[Row]:
    return parse_year_grid(
        data, sheet="1", header_row=2, first_row=3,
        unit_mode="GEL", indicator="median_monthly_earnings",
        dataset_id="median_earnings",
    )


def _earnings_by_region(data: bytes) -> list[Row]:
    return parse_year_grid(
        data, sheet="1", header_row=3, first_row=4,
        unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
        dataset_id="earnings_by_region",
    )


def _earnings_by_activity(data: bytes) -> list[Row]:
    rows: list[Row] = []
    for sheet, prefix in (("NACE1", "nace1"), ("NACE2", "nace2")):
        rows += parse_year_grid(
            data, sheet=sheet, header_row=3, first_row=4,
            unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
            dataset_id="earnings_by_activity", breakdown_prefix=prefix,
        )
    return rows


def _earnings_by_sex(data: bytes) -> list[Row]:
    rows: list[Row] = []
    for sheet in ("NACE1", "NACE2"):
        part = parse_banded_grid(
            data, sheet=sheet, header_row=3, first_row=4,
            unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
            dataset_id="earnings_by_sex",
        )
        prefix = sheet.lower()
        for row in part:
            row.breakdown_code = f"{prefix}.{row.breakdown_code}"
        rows += part
    return rows


def _banded_two_nace(data: bytes, *, dataset_id: str) -> list[Row]:
    """NACE1 + NACE2 sheets, each split into bands whose title sits in column B.

    Earnings by ownership and by business/non-business sector are the same
    shape as earnings by sex, so they reuse the same parser and only differ in
    which sheet prefix the breakdown code carries.
    """
    rows: list[Row] = []
    for sheet in ("NACE1", "NACE2"):
        part = parse_banded_grid(
            data, sheet=sheet, header_row=3, first_row=4,
            unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
            dataset_id=dataset_id,
        )
        prefix = sheet.lower()
        for row in part:
            row.breakdown_code = f"{prefix}.{row.breakdown_code}"
        rows += part
    return rows


def _employment_by_activity(data: bytes) -> list[Row]:
    rows: list[Row] = []
    for sheet, prefix in (("NACE1", "nace1"), ("NACE2", "nace2")):
        rows += parse_year_grid(
            data, sheet=sheet, header_row=3, first_row=4,
            unit_mode="thousand_persons", indicator="employed_persons",
            dataset_id="employment_by_activity", breakdown_prefix=prefix,
        )
    return rows


def _labour_force(data: bytes) -> list[Row]:
    """The labour force sheet mixes head counts and rates in one column of rows.

    `parse_year_grid` stamps one unit across the whole grid, so the three rate
    rows are re-stamped afterwards from their own published label. Guessing by
    magnitude instead would misread a small head count as a percentage.
    """
    rows = parse_year_grid(
        data, sheet="1", header_row=3, first_row=4,
        unit_mode="thousand_persons", indicator="country.georgia",
        dataset_id="labour_force", label_axis="indicator",
    )
    for row in rows:
        if row.indicator_label.rstrip(" .").endswith("percentage"):
            row.unit = "percent"
    return rows


def _earnings_quarterly(data: bytes) -> list[Row]:
    rows: list[Row] = []
    for sheet, prefix in (("NACE1", "nace1"), ("NACE2", "nace2")):
        rows += parse_year_grid(
            data, sheet=sheet, header_row=3, first_row=4,
            unit_mode="currency_era", indicator="avg_monthly_nominal_earnings",
            dataset_id="earnings_quarterly", breakdown_prefix=prefix,
        )
    return rows


def _labour_force_by_region(data: bytes) -> list[Row]:
    return parse_year_banded_matrix(
        data, sheet="1", dataset_id="labour_force_by_region",
        unit_mode="thousand_persons", percent_suffixes=("percentage", "percent"),
    )


def _household_income(data: bytes) -> list[Row]:
    return parse_year_banded_matrix(
        data, sheet="1", dataset_id="household_income",
        unit_mode="GEL_per_household_month", fixed_header_row=2,
    )


def _household_expenditure(data: bytes) -> list[Row]:
    return parse_year_banded_matrix(
        data, sheet="1", dataset_id="household_expenditure",
        unit_mode="GEL_per_household_month", fixed_header_row=2,
    )


def _population(data: bytes) -> list[Row]:
    return parse_geo_year_grid(
        data, sheet="1", header_row=4, first_row=5,
        dataset_id="population", indicator="population_1_january",
        unit="thousand_persons",
    )


def _business_demography(data: bytes) -> list[Row]:
    return parse_indicator_columns_by_year_band(
        data, sheet="By regions", dataset_id="business_demography",
        header_row=4, label_col=2, first_col=3,
        units={
            "active_enterprises": "enterprises",
            "persons_employed_in_the_population_of_active_enterprises": "persons",
            "enterprise_births": "enterprises",
            "enterprise_deaths": "enterprises",
            "enterprise_birth_rate": "percent",
            "enterprise_death_rate": "percent",
        },
    )


def _gender_pay_gap(data: bytes) -> list[Row]:
    return parse_two_level_header(
        data, sheet="1", year_row=2, measure_row=3, first_row=4,
        dataset_id="gender_pay_gap", unit="percent",
        measure_codes={"hourly": "adjusted_gpg_hourly",
                       "monthly": "adjusted_gpg_monthly"},
    )


def _tourism_by_region(data: bytes) -> list[Row]:
    rows = parse_quarter_rows(
        data, sheet="Domestic Tourism", header_row=2, first_row=3,
        dataset_id="tourism_by_region", indicator="domestic_visits",
        unit="thousand_visits",
    )
    rows += parse_quarter_rows(
        data, sheet="Inbound Tourism", header_row=2, first_row=3,
        dataset_id="tourism_by_region", indicator="inbound_visits",
        unit="thousand_visits",
    )
    rows += parse_place_year_grid(
        data, sheet="Hotels", header_row=3, first_row=4,
        dataset_id="tourism_by_region", indicator="hotels",
        # The sheet title says "thousand" but Tbilisi reads 125 in 2006. These
        # are counts of establishments; the word is inherited from the Guests
        # sheet next to it. Recording them as thousands would be a 1000x error.
        unit="enterprises",
    )
    rows += parse_place_year_grid(
        data, sheet="Guests", header_row=3, first_row=4,
        dataset_id="tourism_by_region", indicator="hotel_guests",
        unit="thousand_persons",
    )
    return rows


ADAPTERS: dict[str, Adapter] = {
    a.dataset_id: a for a in [
        Adapter(
            dataset_id="earnings_annual",
            title="Average monthly nominal earnings of employees, annual",
            url="https://geostat.ge/media/77569/01_Earnings_annual.xlsx",
            source_page=WAGES_PAGE,
            parser=_earnings_annual,
            unit_family="currency_era",
            expected_sheets=["1"],
            note=(
                "Headline series, 1970-2025. Columns span four currencies "
                "(Rouble, Coupon, Thousand Coupon, Lari) and are not a "
                "comparable series before 1995."
            ),
        ),
        Adapter(
            dataset_id="median_earnings",
            title="Median earnings of employees by economic activity",
            url="https://geostat.ge/media/73905/Median-Monthly-Earnings.xlsx",
            source_page=WAGES_PAGE,
            parser=_median_earnings,
            unit_family="GEL",
            expected_sheets=["1"],
            note=(
                "Administrative source (Revenue Service), 2018 onward. The gap "
                "between this and the mean is the skew of the wage distribution."
            ),
        ),
        Adapter(
            dataset_id="earnings_by_region",
            title="Average monthly nominal earnings by region",
            url="https://geostat.ge/media/73904/13_Earnings-by-regions_annual.xlsx",
            source_page=WAGES_PAGE,
            parser=_earnings_by_region,
            unit_family="currency_era",
            expected_sheets=["1"],
            note=(
                "Enterprises are counted at the location of the head office, "
                "which inflates Tbilisi relative to where work is done."
            ),
        ),
        Adapter(
            dataset_id="earnings_by_sex",
            title="Average monthly nominal earnings by economic activity and sex",
            url="https://geostat.ge/media/73901/03_Earnings-by-sex_annual.xlsx",
            source_page=WAGES_PAGE,
            parser=_earnings_by_sex,
            unit_family="currency_era",
            expected_sheets=["NACE1", "NACE2"],
            note=(
                "Two activity classifications. NACE rev.1 covers 1999-2019, "
                "NACE rev.2 covers 2014-2024; they are not splice-compatible."
            ),
        ),
        Adapter(
            dataset_id="earnings_by_activity",
            title="Average monthly nominal earnings by economic activity",
            url="https://geostat.ge/media/73900/02_Earnings-by-activity_annual.xlsx",
            source_page=WAGES_PAGE,
            parser=_earnings_by_activity,
            unit_family="currency_era",
            expected_sheets=["NACE1", "NACE2"],
            note=(
                "The 2006 column carries a methodology footnote: the drop in "
                "Financial intermediation is a coverage change, not a wage fall."
            ),
        ),
        Adapter(
            dataset_id="cpi_2010_base",
            title="Consumer price index, 2010 average = 100",
            url="https://geostat.ge/media/81846/consumer-price-index-2010%3D100-%2810%29.xlsx",
            source_page=INFLATION_PAGE,
            parser=lambda d: parse_cpi_monthly(d, dataset_id="cpi_2010_base"),
            unit_family="index_2010_100",
            expected_sheets=["Georgia "],
            period_grain="monthly",
            note=(
                "Monthly index for Georgia and five cities. GeoStats derives an "
                "annual average only from years with all twelve months present."
            ),
        ),
        Adapter(
            dataset_id="cpi_yoy",
            title="Consumer price index vs the same month of the previous year",
            url="https://geostat.ge/media/81847/consumer-price-index-to-the-same-month-of-previous-year.xlsx",
            source_page=INFLATION_PAGE,
            parser=lambda d: parse_cpi_yoy(d, dataset_id="cpi_yoy"),
            unit_family="index_prev_year_100",
            expected_sheets=["Georgia"],
            period_grain="monthly",
            note=(
                "Headline annual inflation rate as published monthly. Only the "
                "Total line is ingested; the COICOP tree is out of scope."
            ),
        ),
        Adapter(
            dataset_id="basket_weights",
            title="Consumer basket weights",
            url="https://geostat.ge/media/76662/Consumer-basket-weights.xlsx",
            source_page=INFLATION_PAGE,
            parser=lambda d: parse_basket_weights(d, dataset_id="basket_weights"),
            unit_family="share_of_1",
            expected_sheets=["Weights"],
            note=(
                "The weights behind the CPI. Published as shares summing to 1 "
                "even though the sheet title says percent."
            ),
        ),
        Adapter(
            dataset_id="labour_force",
            title="Labour force indicators, annual",
            url="https://geostat.ge/media/78735/01-Labour-Force-Indicators.xlsx",
            source_page=EMPLOYMENT_PAGE,
            parser=_labour_force,
            unit_family="thousand_persons",
            expected_sheets=["1"],
            note=(
                "Labour Force Survey, 1998 onward, on the ILO definition and "
                "covering the population aged 15 and over: a person is "
                "unemployed only if they are out of work, available for work "
                "and actively looking for it, so someone who has stopped "
                "looking is outside the labour force rather than unemployed. "
                "Head counts and rates share one sheet, so the unit is taken "
                "from each row's own label. The series breaks between 2009 and "
                "2010: Geostat adopted the ICLS-19 standard in 2020, which "
                "puts subsistence farmers outside employment, and recalculated "
                "2010 onward while leaving 1998-2009 on the old basis. "
                "Employment falls by 443 thousand across that boundary for "
                "definitional reasons."
            ),
        ),
        Adapter(
            dataset_id="employment_by_activity",
            title="Employed persons by economic activity",
            url="https://geostat.ge/media/78741/07-Employment-by-branch-NACE_Eng.xlsx",
            source_page=EMPLOYMENT_PAGE,
            parser=_employment_by_activity,
            unit_family="thousand_persons",
            expected_sheets=["NACE1", "NACE2"],
            note=(
                "Head counts, not earnings: the denominator behind the wage "
                "series. NACE rev.1 stops at 2020 and rev.2 starts at 2017; "
                "they overlap but are not splice-compatible."
            ),
        ),
        Adapter(
            dataset_id="earnings_by_ownership",
            title="Average monthly nominal earnings by ownership and activity",
            url="https://geostat.ge/media/73903/10-Earnings-ownership_%28annual%29.xlsx",
            source_page=WAGES_PAGE,
            parser=lambda d: _banded_two_nace(d, dataset_id="earnings_by_ownership"),
            unit_family="currency_era",
            expected_sheets=["NACE1", "NACE2"],
            note=(
                "Public against non-public sector. The gap is a composition "
                "effect as much as a pay policy: the two sectors do not employ "
                "the same mix of occupations."
            ),
        ),
        Adapter(
            dataset_id="earnings_business_sector",
            title="Average monthly nominal earnings by business/non-business sector",
            url=(
                "https://geostat.ge/media/73902/"
                "06_Earnings-by-Business-Non-business--sectors_annual.xlsx"
            ),
            source_page=WAGES_PAGE,
            parser=lambda d: _banded_two_nace(
                d, dataset_id="earnings_business_sector"
            ),
            unit_family="currency_era",
            expected_sheets=["NACE1", "NACE2"],
            note=(
                "Business sector is the enterprise survey; non-business is "
                "budget-funded organisations. Different collection instruments, "
                "so the levels are not strictly comparable."
            ),
        ),
        Adapter(
            dataset_id="earnings_quarterly",
            title="Average monthly nominal earnings by economic activity, quarterly",
            url="https://geostat.ge/media/80118/04-earnings-quarterly.xlsx",
            source_page=WAGES_PAGE,
            parser=_earnings_quarterly,
            unit_family="currency_era",
            expected_sheets=["NACE1", "NACE2"],
            period_grain="quarterly",
            note=(
                "The same earnings measured four times a year. Georgian wages "
                "carry a strong fourth-quarter bonus effect, so Q4 against Q3 "
                "is seasonality, not growth; compare a quarter with the same "
                "quarter a year earlier."
            ),
        ),
        Adapter(
            dataset_id="labour_force_by_region",
            title="Labour force indicators by region",
            url="https://geostat.ge/media/78739/05-Labour-Force-Indicators-by-regions.xlsx",
            source_page=EMPLOYMENT_PAGE,
            parser=_labour_force_by_region,
            unit_family="thousand_persons",
            expected_sheets=["1"],
            note=(
                "Published one block per year, and the set of regions changes: "
                "the early years fold several into 'The remaining regions'. "
                "Unlike the earnings file this is a household survey, so a "
                "person is counted where they live, not where their employer "
                "is registered. Carries the same 2009/2010 definition break as "
                "the national series, and stops publishing the hired / "
                "self-employed split for 2010-2019."
            ),
        ),
        Adapter(
            dataset_id="household_income",
            title="Average monthly household income by region",
            url=(
                "https://geostat.ge/media/79265/"
                "106_Distribution-of-average-monthly-incomes-per-household-by-regions.xlsx"
            ),
            source_page=HOUSEHOLD_INCOME_PAGE,
            parser=_household_income,
            unit_family="GEL_per_household_month",
            expected_sheets=["1"],
            note=(
                "Self-reported in the Household Incomes and Expenditures Survey "
                "and known to understate income. It is not comparable with the "
                "earnings series: that one counts a job, this one counts a "
                "household and includes pensions, transfers and own produce."
            ),
        ),
        Adapter(
            dataset_id="household_expenditure",
            title="Average monthly household expenditure by region",
            url=(
                "https://geostat.ge/media/79283/"
                "206_Distribution-of-average-monthly-expenditures-per-household-by-regions.xlsx"
            ),
            source_page=HOUSEHOLD_EXPENDITURE_PAGE,
            parser=_household_expenditure,
            unit_family="GEL_per_household_month",
            expected_sheets=["1"],
            note=(
                "Same survey as household income. Reported expenditure usually "
                "exceeds reported income, which is a known survey artefact: "
                "the difference is not saving or dissaving and must not be "
                "presented as one."
            ),
        ),
        Adapter(
            dataset_id="population",
            title="Population by region and municipality, 1 January",
            url="https://geostat.ge/media/78356/01-population-by-self-governed-unit.xlsx",
            source_page=POPULATION_PAGE,
            parser=_population,
            unit_family="thousand_persons",
            expected_sheets=["1"],
            note=(
                "Regions and their municipalities share one column, "
                "distinguished only by the cell indent. Rebased on the 2024 "
                "census, so the whole series was revised. Occupied territories "
                "are carried at their last enumerated figures."
            ),
        ),
        Adapter(
            dataset_id="business_demography",
            title="Enterprise births, deaths and employment by region",
            url="https://geostat.ge/media/75003/Main-results.xlsx",
            source_page=BUSINESS_DEMOGRAPHY_PAGE,
            parser=_business_demography,
            unit_family="enterprises",
            expected_sheets=["By regions"],
            note=(
                "Registrations, not economic activity: an individual "
                "entrepreneur registering counts the same as a factory opening. "
                "Enterprises are attributed to their registered address, which "
                "concentrates them in Tbilisi for the same reason the earnings "
                "series does."
            ),
        ),
        Adapter(
            dataset_id="gender_pay_gap",
            title="Adjusted gender pay gap, hourly and monthly",
            url="https://geostat.ge/media/74461/GPG_Eng.xlsx",
            source_page=WAGES_PAGE,
            parser=_gender_pay_gap,
            unit_family="percent",
            expected_sheets=["1"],
            note=(
                "Adjusted for observable characteristics, so it is the gap "
                "remaining after composition is accounted for, not the raw "
                "difference in average pay. The hourly and monthly figures "
                "differ because women work fewer paid hours; the monthly gap "
                "is the larger of the two every published year."
            ),
        ),
        Adapter(
            dataset_id="tourism_by_region",
            title="Tourism visits, hotels and guests by region",
            url="https://geostat.ge/media/81807/tourism_regions.xlsx",
            source_page=REGIONAL_PAGE,
            parser=_tourism_by_region,
            unit_family="thousand_visits",
            expected_sheets=["Domestic Tourism", "Inbound Tourism", "Hotels",
                             "Guests"],
            period_grain="mixed",
            note=(
                "Visits are quarterly survey estimates; hotels and guests are "
                "annual declared administrative counts. The 2020 and 2021 "
                "collapse is the pandemic and is left in place rather than "
                "smoothed."
            ),
        ),
    ]
}
